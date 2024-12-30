import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import os

def get_url_data(url, description='没有说明'):
    print(f"本次请求为：【{description}】网络请求--------------")
    header = {"user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36 Edg/107.0.1418.56"}
    response = requests.get(url=url, headers=header)
    if response.status_code == 200:
        return response.content
    else:
        print(f"网络访问失败-请检查网站信息: {url}")
        return None


def parse_page(data, worksheet, year, month):
    html = data.decode('utf-8')
    page_soup = BeautifulSoup(html, 'html.parser')
    weather_datas = page_soup.find('ul', attrs={'class': 'tian_two'}).find_all('li')
    weather_list = []
    for weather_data in weather_datas:
        datas = weather_data.find_all('div', attrs={'class': 'tian_twoa'})
        for i in datas:
            weather_list.append(i.string.strip())  # 使用strip()去除可能的空白字符

    # 构建年月字符串
    year_month = f"{year}{month:02d}"

    header_row = ['年月', '月平均高温', '月平均低温', '月最高温', '月最低温', '月空气质量指数', '空气最好指数', '空气最差指数']
    worksheet.append(header_row)

    data_row = [year_month] + weather_list
    worksheet.append(data_row)

def main():
    file_path = '北京18-21年每月空气质量和最高最低温.xlsx'
    if not os.path.exists(file_path):
        wk = Workbook()
        wb = wk.active
        wb.title = "北京每月空气质量指数和最高最低温"
        wk.save(filename=file_path)

    wk = load_workbook(filename=file_path)
    wb = wk['北京每月空气质量指数和最高最低温']

    for year in range(2018, 2022):
        for month in range(1, 13):
            if month < 10:
                url = f'https://lishi.tianqi.com/beijing/{year}0{month}.html'
            else:
                url = f'https://lishi.tianqi.com/beijing/{year}{month}.html'
            print(f"爬取第{year}年第{month}月数据")
            page_data = get_url_data(url=url, description=f'爬取第{year}年第{month}月数据')
            if page_data:
                parse_page(page_data, wb, year, month)

    wk.save(filename=file_path)

if __name__ == '__main__':
    main()
